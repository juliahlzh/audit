import json
from pathlib import Path
import unittest
from datetime import date
from io import BytesIO
from unittest.mock import patch

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app.auth import hash_password
from app.database import Base, get_db
from app.main import app
from app.models import BranchInput, MatchingResult, User
from app.services.analytics import build_monitoring_context
from app.services.organization import ORGANIZATION_CODE_ROWS, ORGANIZATION_ROWS, REGIONAL_ACCOUNTS, resolve_location
from app.services.reports import build_ranked_excel_report
from app.services.sample_loader import load_sample_file


class MonitoringFeatureTests(unittest.TestCase):
    def setUp(self):
        self.engine = create_engine(
            "sqlite:///:memory:", connect_args={"check_same_thread": False}, poolclass=StaticPool
        )
        Base.metadata.create_all(bind=self.engine)
        self.Session = sessionmaker(bind=self.engine)
        self.db = self.Session()
        self.admin = User(username="admin2", full_name="Admin Monitor", password_hash=hash_password("secret"), role="admin")
        self.jabar = User(username="jabar2", full_name="Auditor Jabar", password_hash=hash_password("secret"), role="viewer", region="Jawa Barat")
        self.db.add_all([self.admin, self.jabar])
        self.db.flush()

        self.jbr_result = self._add_result(
            region="Jawa Barat", location="Bandung", invoice="JBR-1", score=10,
            follow_up="RESOLVED", code="amount_mismatch", name="Jumlah biaya tidak sesuai jumlah setor",
        )
        self.jtg_result = self._add_result(
            region="Jawa Tengah", location="Semarang Uji", invoice="JTG-1", score=4,
            follow_up="OPEN", code="late_input_transfer", name="Keterlambatan input data setor transfer",
        )
        self.db.commit()

        def override_get_db():
            yield self.db

        app.dependency_overrides[get_db] = override_get_db
        self.client = TestClient(app)

    def tearDown(self):
        app.dependency_overrides.clear()
        self.db.close()
        Base.metadata.drop_all(bind=self.engine)
        self.engine.dispose()

    def _add_result(self, *, region, location, invoice, score, follow_up, code, name, transaction_date=None):
        branch = BranchInput(
            transaction_date=transaction_date or date(2026, 6, 10), region=region, area=f"Area {location}", branch_name=location,
            customer_name="DATA SINTETIS", amount_should_pay=1_000_000,
            amount_input_branch=900_000, payment_method="transfer", invoice_code=invoice,
            data_type="UJI",
        )
        self.db.add(branch)
        self.db.flush()
        result = MatchingResult(
            branch_input_id=branch.id, status="UNMATCHED" if score > 7 else "NEED REVIEW",
            risk_score=score, risk_level="High Alert" if score > 7 else "Medium",
            triggered_rules=json.dumps([{"code": code, "name": name, "score": score}]),
            follow_up_status=follow_up,
        )
        self.db.add(result)
        self.db.flush()
        return result

    def _login(self, username):
        response = self.client.post("/login", data={"username": username, "password": "secret"}, follow_redirects=False)
        self.assertEqual(response.status_code, 303)

    def test_navigation_contains_dashboard_info_and_reports(self):
        self._login("admin2")
        response = self.client.get("/dashboard")

        self.assertEqual(response.status_code, 200)
        self.assertIn('href="/dashboard"', response.text)
        self.assertIn('href="/info"', response.text)
        self.assertIn('href="/reports"', response.text)
        self.assertIn('href="/branch-inputs"', response.text)
        self.assertIn('href="/alerts"', response.text)
        self.assertIn("Upload Data", response.text)
        self.assertIn("Dashboard Pusat FEWS", response.text)
        self.assertNotIn('action="/branch-inputs/upload"', response.text)
        self.assertNotIn("legacy-info", response.text)

    def test_info_page_contains_complete_legacy_dashboard_information(self):
        self._login("admin2")
        response = self.client.get("/info")

        self.assertEqual(response.status_code, 200)
        for heading in [
            "Indikator yang Dites",
            "Aturan Skor Risiko",
            "Tren Approval Terbaru",
            "Indikator Paling Sering",
            "Transaksi Prioritas Investigasi",
            "Aktivitas Sistem",
        ]:
            self.assertIn(heading, response.text)
        self.assertIn("H+2 hari kerja dari tanggal bank", response.text)
        self.assertIn("Upload Data", response.text)

    def test_region_account_cannot_open_central_info(self):
        self._add_result(
            region="Jawa Tengah", location="Solo", invoice="JTG-HIGH", score=9,
            follow_up="OPEN", code="amount_mismatch", name="Jumlah biaya tidak sesuai jumlah setor",
        )
        self.db.commit()
        self._login("jabar2")

        response = self.client.get("/info")

        self.assertEqual(response.status_code, 403)

    def test_region_account_has_read_only_alert_center_and_scoped_data(self):
        self._login("jabar2")
        dashboard = self.client.get("/dashboard")
        alerts = self.client.get("/alerts")
        upload_page = self.client.get("/branch-inputs")
        upload_post = self.client.post(
            "/branch-inputs/upload",
            files={"excel_file": ("data.csv", b"invalid", "text/csv")},
        )
        own_denied = self.client.post(f"/reports/{self.jbr_result.id}/verify", data={"notes": "Tidak boleh"})

        self.assertIn('href="/alerts"', dashboard.text)
        self.assertIn('aria-label="Tutup navigasi"', dashboard.text)
        self.assertNotIn('href="/info"', dashboard.text)
        self.assertNotIn('href="/branch-inputs"', dashboard.text)
        self.assertNotIn('action="/branch-inputs/upload"', dashboard.text)
        self.assertIn("Dashboard Wilayah Jawa Barat", dashboard.text)
        self.assertIn("Jawa Tengah", dashboard.text)
        self.assertNotIn("JTG-1", dashboard.text)
        self.assertIn("JBR-1", alerts.text)
        self.assertNotIn("JTG-1", alerts.text)
        self.assertIn("Mode lihat saja", alerts.text)
        self.assertEqual(upload_page.status_code, 403)
        self.assertEqual(upload_post.status_code, 403)
        self.assertEqual(own_denied.status_code, 403)

    def test_organization_master_matches_source_deck(self):
        self.assertEqual(len(REGIONAL_ACCOUNTS), 15)
        self.assertEqual(len(ORGANIZATION_ROWS), 166)
        self.assertEqual(len({area for _, area, _ in ORGANIZATION_ROWS}), 41)
        self.assertEqual(len(ORGANIZATION_CODE_ROWS), 166)
        self.assertEqual(len({code for code, _, _, _ in ORGANIZATION_CODE_ROWS}), 166)

    def test_sil_location_code_maps_name_region_and_area(self):
        self.assertEqual(
            resolve_location("278"),
            ("278", "Merduati", "Sumatera Bagian Utara", "Area Aceh"),
        )
        self.assertEqual(
            resolve_location(223),
            ("223", "Graha Mustika Media", "Bekasi Plus", "Area Kabupaten Bogor"),
        )

        branch = BranchInput(
            transaction_date=date(2026, 7, 1), branch_name="278", customer_name="DATA UJI",
            amount_should_pay=100_000, amount_input_branch=100_000, payment_method="transfer",
            invoice_code="SIL-278",
        )
        self.db.add(branch)
        self.db.flush()
        self.assertEqual(branch.location_code, "278")
        self.assertEqual(branch.branch_name, "Merduati")
        self.assertEqual(branch.region, "Sumatera Bagian Utara")
        self.assertEqual(branch.area, "Area Aceh")

    def test_region_and_verification_filters_change_dashboard(self):
        self._login("admin2")
        response = self.client.get("/dashboard?region=Jawa%20Barat&month=2026-06&verification=sudah")

        self.assertIn("JBR-1", response.text)
        self.assertNotIn("JTG-1", response.text)
        self.assertIn("Sudah Diverifikasi", response.text)

    def test_dashboard_accepts_sil_code_as_location_filter(self):
        mapped = self._add_result(
            region="Belum Dipetakan", location="278", invoice="SIL-FILTER-278", score=8,
            follow_up="OPEN", code="amount_mismatch", name="Jumlah biaya tidak sesuai jumlah setor",
        )
        self.db.commit()
        self._login("admin2")

        response = self.client.get("/dashboard?location=278&month=2026-06")

        self.assertEqual(response.status_code, 200)
        self.assertIn("SIL-FILTER-278", response.text)
        self.assertEqual(mapped.branch_input.location_code, "278")
        self.assertEqual(mapped.branch_input.area, "Area Aceh")
        self.assertIn('value="Merduati" selected', response.text)

    def test_weekly_period_filters_dashboard_and_shows_complete_fews_sections(self):
        self._add_result(
            region="Jawa Barat", location="Bogor", invoice="JBR-W25", score=8,
            follow_up="OPEN", code="amount_mismatch", name="Jumlah biaya tidak sesuai jumlah setor",
            transaction_date=date(2026, 6, 17),
        )
        self.db.commit()
        self._login("admin2")

        response = self.client.get(
            "/dashboard?region=Jawa%20Barat&period_type=mingguan&week=2026-W24"
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn("JBR-1", response.text)
        self.assertNotIn("JBR-W25", response.text)
        for heading in [
            "Temuan per Wilayah", "Temuan per Lokasi",
            "Penyebab Utama berdasarkan Lokasi", "Detail Temuan Terbaru",
        ]:
            self.assertIn(heading, response.text)
        self.assertEqual(response.text.count('data-chart-type="bar"'), 2)
        self.assertIn('data-chart-action="in"', response.text)
        self.assertIn('data-chart-action="out"', response.text)
        self.assertIn('data-chart-action="reset"', response.text)
        self.assertNotIn("Upload Excel ke FEWS", response.text)
        self.assertIn('value="mingguan" selected', response.text)
        self.assertIn('value="2026-W24"', response.text)

    def test_dashboard_bar_data_contains_every_filtered_location_without_top_ten_limit(self):
        for index in range(12):
            self._add_result(
                region="Jawa Barat",
                location=f"Lokasi Uji {index + 1:02d}",
                invoice=f"BAR-{index + 1:02d}",
                score=(index % 9) + 1,
                follow_up="OPEN",
                code="amount_mismatch",
                name="Jumlah biaya tidak sesuai jumlah setor",
            )
        self.db.commit()

        context = build_monitoring_context(
            self.db,
            self.admin,
            {
                "region": "Jawa Barat",
                "area": "",
                "location": "",
                "period_type": "bulanan",
                "month": "2026-06",
                "week": "",
                "indicator": "",
                "verification": "",
            },
        )

        self.assertEqual(
            len(context["dashboard_location_bar_rows"]),
            len(context["location_rows"]),
        )
        self.assertEqual(len(context["dashboard_location_bar_rows"]), 13)
        self.assertIn(
            "Lokasi Uji 12",
            [row["name"] for row in context["dashboard_location_bar_rows"]],
        )
        self.assertTrue(all(row["value"] >= 1 for row in context["dashboard_location_bar_rows"]))

    def test_area_filter_limits_results(self):
        self._login("admin2")
        response = self.client.get("/dashboard?area=Area%20Bandung&month=2026-06")

        self.assertIn("JBR-1", response.text)
        self.assertNotIn("JTG-1", response.text)

    def test_region_user_cannot_see_or_verify_other_region(self):
        self._login("jabar2")
        response = self.client.get("/reports?region=Jawa%20Tengah")
        denied = self.client.post(f"/reports/{self.jtg_result.id}/verify", data={"notes": "Tidak boleh"})

        self.assertIn("JBR-1", response.text)
        self.assertNotIn("JTG-1", response.text)
        self.assertEqual(denied.status_code, 403)

    def test_ranking_and_excel_are_sorted_from_most_severe(self):
        context = build_monitoring_context(
            self.db, self.admin,
            {"region": "", "area": "", "location": "", "month": "2026-06", "indicator": "", "verification": ""},
        )
        self.assertEqual([row["name"] for row in context["location_rows"]], ["Bandung", "Semarang Uji"])

        payload = build_ranked_excel_report(context["location_rows"], context["filters"])
        sheet = load_workbook(BytesIO(payload))["Ranking Lokasi"]
        self.assertEqual(sheet["A5"].value, "Peringkat")
        self.assertEqual(sheet["E6"].value, "Bandung")
        self.assertEqual(sheet["E7"].value, "Semarang Uji")
        self.assertEqual(sheet.tables["RankingLokasiFEWS"].ref, "A5:O7")

        complete_payload = build_ranked_excel_report(
            context["location_rows"], context["filters"], detail_rows=context["detail_rows"],
            trend=context["trend"], indicator_rows=context["indicator_rows"],
        )
        workbook = load_workbook(BytesIO(complete_payload))
        self.assertEqual(
            workbook.sheetnames,
            ["Ranking Lokasi", "10 Risiko Tertinggi", "10 Terendah", "Detail Temuan", "Tren Periode", "Grafik Indikator", "Grafik Lokasi"],
        )
        self.assertEqual(workbook["Detail Temuan"]["A1"].value, "ID Unix")
        self.assertEqual(workbook["Detail Temuan"]["E1"].value, "Kode Lokasi")
        self.assertEqual(workbook["Detail Temuan"]["H1"].value, "Jumlah Kesalahan")
        self.assertIsNone(workbook["Ranking Lokasi"].auto_filter.ref)
        self.assertIsNone(workbook["10 Risiko Tertinggi"].auto_filter.ref)
        self.assertIsNone(workbook["Detail Temuan"].auto_filter.ref)
        self.assertEqual(len(workbook["Tren Periode"]._charts), 1)
        self.assertEqual(len(workbook["Grafik Indikator"]._charts), 1)
        self.assertEqual(len(workbook["Grafik Lokasi"]._charts), 1)
        self.assertIsNotNone(workbook["Tren Periode"]._charts[0].series[0].cat.strRef)
        self.assertIsNotNone(workbook["Grafik Indikator"]._charts[0].series[0].cat.strRef)
        self.assertIsNotNone(workbook["Grafik Lokasi"]._charts[0].series[0].cat.strRef)

    def test_region_account_exports_excel_only_for_its_own_region(self):
        self._login("jabar2")
        response = self.client.get("/reports/excel?region=Jawa%20Tengah&month=2026-06")

        self.assertEqual(response.status_code, 200)
        self.assertIn("fews_jawa_barat_ranking_lokasi.xlsx", response.headers["content-disposition"])
        sheet = load_workbook(BytesIO(response.content))["Ranking Lokasi"]
        exported_regions = {sheet.cell(row=row, column=2).value for row in range(6, sheet.max_row + 1)}
        self.assertEqual(exported_regions, {"Jawa Barat"})

    def test_excel_export_escapes_uploaded_formula_text(self):
        self.jbr_result.branch_input.invoice_code = "=HYPERLINK(\"https://invalid.example\")"
        self.db.commit()
        context = build_monitoring_context(
            self.db, self.admin,
            {"region": "Jawa Barat", "area": "", "location": "", "month": "2026-06", "indicator": "", "verification": ""},
        )

        payload = build_ranked_excel_report(
            context["location_rows"], context["filters"], detail_rows=context["detail_rows"],
        )
        cell = load_workbook(BytesIO(payload))["Detail Temuan"]["A2"]

        self.assertEqual(cell.data_type, "s")
        self.assertTrue(cell.value.startswith("'="))

    def test_admin_must_select_region_before_excel_or_pdf_export(self):
        self._login("admin2")

        excel = self.client.get("/reports/excel")
        pdf = self.client.get("/reports/pdf")

        self.assertEqual(excel.status_code, 400)
        self.assertEqual(pdf.status_code, 400)
        self.assertIn("Pilih satu wilayah", excel.json()["detail"])
        self.assertIn("Pilih satu wilayah", pdf.json()["detail"])

    def test_pdf_export_uses_active_filters_and_locked_region(self):
        self._login("jabar2")
        with patch("app.main.build_pdf_report", return_value=b"%PDF-FAKE") as pdf_builder:
            response = self.client.get(
                "/reports/pdf?region=Jawa%20Tengah&month=2026-06&verification=sudah"
            )

        self.assertEqual(response.status_code, 200)
        self.assertIn("fews_jawa_barat_laporan.pdf", response.headers["content-disposition"])
        rows, region = pdf_builder.call_args.args
        self.assertEqual(region, "Jawa Barat")
        self.assertEqual([row.branch_input.invoice_code for row in rows], ["JBR-1"])

    def test_report_page_offers_excel_and_pdf_exports(self):
        self._login("admin2")
        response = self.client.get("/reports?region=Jawa%20Barat")

        self.assertEqual(response.status_code, 200)
        self.assertIn('action="/reports/excel"', response.text)
        self.assertIn('formaction="/reports/pdf"', response.text)
        self.assertIn("Tabel Detail FEWS", response.text)
        self.assertIn("10 Risiko Tertinggi", response.text)

    def test_real_pdf_export_contains_complete_report(self):
        self._login("admin2")
        response = self.client.get("/reports/pdf?region=Jawa%20Barat&period_type=bulanan&month=2026-06")

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.content.startswith(b"%PDF"))
        self.assertGreater(len(response.content), 3000)

    def test_admin_can_mark_finding_verified(self):
        self._login("admin2")
        response = self.client.post(
            f"/reports/{self.jtg_result.id}/verify", data={"notes": "Bukti sudah diperiksa"}, follow_redirects=False
        )
        self.db.refresh(self.jtg_result)

        self.assertEqual(response.status_code, 303)
        self.assertEqual(self.jtg_result.follow_up_status, "RESOLVED")
        self.assertEqual(self.jtg_result.follow_up_notes, "Bukti sudah diperiksa")

    def test_sample_loader_is_idempotent_and_labels_synthetic_data(self):
        path = Path(__file__).resolve().parents[1] / "sample_data" / "fews_uji.csv"
        first = load_sample_file(self.db, path)
        second = load_sample_file(self.db, path)

        self.assertEqual(first["inserted"], 8)
        self.assertEqual(second["inserted"], 0)
        loaded = self.db.query(BranchInput).filter(BranchInput.invoice_code.like("UJI-%")).all()
        self.assertEqual(len(loaded), 8)
        self.assertTrue(all(row.data_type == "UJI" for row in loaded))
        self.assertTrue(all("SINTETIS" in row.customer_name for row in loaded))
        self.assertTrue(all(row.location_code for row in loaded))
        self.assertTrue(all(row.area != "Belum Dipetakan" for row in loaded))


if __name__ == "__main__":
    unittest.main()
