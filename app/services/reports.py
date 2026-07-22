from io import BytesIO
import json
from collections import Counter, defaultdict

DAY_NAMES_ID = {
    0: "Senin",
    1: "Selasa",
    2: "Rabu",
    3: "Kamis",
    4: "Jumat",
    5: "Sabtu",
    6: "Minggu",
}


def _excel_safe(value):
    """Keep uploaded text from being interpreted as a spreadsheet formula."""
    if isinstance(value, str) and value.lstrip().startswith(("=", "+", "-", "@")):
        return f"'{value}"
    return value


def _excel_safe_row(values) -> list:
    return [_excel_safe(value) for value in values]


def _date_label(value) -> str:
    if not value:
        return "-"
    return f"{DAY_NAMES_ID[value.weekday()]}, {value:%d/%m/%Y}"


def _rule_names(txn) -> list[str]:
    try:
        rules = json.loads(txn.triggered_rules or "[]")
    except (TypeError, json.JSONDecodeError):
        rules = []
    names = [rule.get("name") or rule.get("code") for rule in rules if rule.get("name") or rule.get("code")]
    if names:
        return names
    if txn.mismatch_type:
        return [part.strip() for part in txn.mismatch_type.split(";") if part.strip()]
    return []


def summarize_by_location(transactions):
    grouped = defaultdict(
        lambda: {
            "location": "-",
            "total": 0,
            "high": 0,
            "medium": 0,
            "low": 0,
            "score_sum": 0,
            "max_score": 0,
            "indicator_counts": Counter(),
            "follow_up_counts": Counter(),
            "latest_date": None,
        }
    )

    for txn in transactions:
        branch = txn.branch_input.branch_name if txn.branch_input else "-"
        row = grouped[branch]
        row["location"] = branch
        row["total"] += 1
        row["score_sum"] += txn.risk_score or 0
        row["max_score"] = max(row["max_score"], txn.risk_score or 0)

        if (txn.risk_score or 0) > 7:
            row["high"] += 1
        elif 4 <= (txn.risk_score or 0) <= 7:
            row["medium"] += 1
        else:
            row["low"] += 1

        for name in _rule_names(txn):
            row["indicator_counts"][name] += 1

        row["follow_up_counts"][txn.follow_up_status or "OPEN"] += 1

        if txn.branch_input and txn.branch_input.transaction_date:
            current_date = txn.branch_input.transaction_date
            if row["latest_date"] is None or current_date > row["latest_date"]:
                row["latest_date"] = current_date

    summaries = []
    for row in grouped.values():
        avg_score = (row["score_sum"] / row["total"]) if row["total"] else 0
        indicator_summary = "; ".join(
            f"{name} ({count})" for name, count in row["indicator_counts"].most_common(5)
        ) or "Tidak ada indikator fraud"
        follow_up_summary = "; ".join(
            f"{status} ({count})" for status, count in row["follow_up_counts"].most_common()
        ) or "-"
        if row["high"]:
            risk_label = "Tinggi"
        elif row["medium"]:
            risk_label = "Sedang"
        else:
            risk_label = "Rendah"

        summaries.append(
            {
                "location": row["location"],
                "total": row["total"],
                "high": row["high"],
                "medium": row["medium"],
                "low": row["low"],
                "max_score": row["max_score"],
                "avg_score": round(avg_score, 1),
                "risk_label": risk_label,
                "indicator_summary": indicator_summary,
                "follow_up_summary": follow_up_summary,
                "latest_date": row["latest_date"],
            }
        )

    return sorted(summaries, key=lambda item: (-item["high"], -item["medium"], -item["total"], item["location"]))


def build_excel_report(transactions):
    import pandas as pd

    rows = [
            {
                "Lokasi/Cabang": row["location"],
                "Total Transaksi": row["total"],
                "High Alert": row["high"],
                "Need Review": row["medium"],
                "Normal/Rendah": row["low"],
                "Skor Tertinggi": row["max_score"],
                "Rata-rata Skor": row["avg_score"],
                "Risiko Lokasi": row["risk_label"],
                "Indikator Utama": row["indicator_summary"],
                "Status Tindak Lanjut": row["follow_up_summary"],
                "Tanggal Terakhir": _date_label(row["latest_date"]),
            }
            for row in summarize_by_location(transactions)
    ]
    df = pd.DataFrame(rows)
    if not df.empty:
        df = df.map(_excel_safe)
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Ringkasan Lokasi")
    return output.getvalue()


def _verification_label(status: str | None) -> str:
    return "Sudah Diverifikasi" if (status or "").upper() == "RESOLVED" else "Belum Diverifikasi"


def _detail_row(result) -> list:
    branch = result.branch_input
    names = _rule_names(result)
    return _excel_safe_row([
        branch.invoice_code,
        branch.transaction_date.strftime("%d/%m/%Y"),
        branch.region,
        branch.area,
        branch.location_code or "",
        branch.branch_name,
        "; ".join(names) or "Tidak ada indikator",
        len(names),
        result.risk_score or 0,
        branch.data_type,
        _verification_label(result.follow_up_status),
    ])


def build_ranked_excel_report(
    location_rows,
    filters,
    *,
    detail_rows=None,
    trend=None,
    indicator_rows=None,
):
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.chart.data_source import AxDataSource, StrRef
    from openpyxl.chart.label import DataLabelList
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    detail_rows = detail_rows or []
    trend = trend or {"labels": [], "series": []}
    indicator_rows = indicator_rows or []
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Ranking Lokasi"
    sheet.append(["Laporan Ranking FEWS"])
    sheet["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="0B2B4C")
    sheet.merge_cells("A1:O1")
    sheet.append(["Dibuat", datetime.now().strftime("%d/%m/%Y %H:%M")])
    sheet.append(_excel_safe_row([
        "Filter",
        "; ".join(
            f"{label}: {filters.get(key) or 'Semua'}"
            for key, label in [
                ("region", "Wilayah"), ("area", "Area"), ("location", "Lokasi"),
                ("period_type", "Jenis Periode"), ("month", "Bulan"), ("week", "Minggu"),
                ("indicator", "Kesalahan"), ("verification", "Verifikasi"),
            ]
        ),
    ]))
    sheet.append([])
    headers = [
        "Peringkat", "Wilayah", "Area", "Kode Lokasi", "Lokasi", "Total Temuan", "High Alert", "Need Review",
        "Risiko Rendah", "Total Skor", "Skor Tertinggi", "Rata-rata Skor", "Tingkat Risiko",
        "Sudah Diverifikasi", "Belum Diverifikasi",
    ]
    sheet.append(headers)
    for row in location_rows:
        sheet.append(_excel_safe_row([
            row["rank"], row["region"], row["area"], row.get("code", ""), row["name"], row["total"], row["high"], row["medium"],
            row["low"], row["score_total"], row["max_score"], row["avg_score"], row["risk_label"],
            row["verified"], row["unverified"],
        ]))

    header_row = 5
    if location_rows:
        table = Table(displayName="RankingLokasiFEWS", ref=f"A{header_row}:O{sheet.max_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
            showRowStripes=True, showColumnStripes=False,
        )
        sheet.add_table(table)
    for cell in sheet[header_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="145DA0")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.freeze_panes = "A6"
    if not location_rows:
        sheet.auto_filter.ref = f"A{header_row}:O{header_row}"
    widths = [12, 22, 24, 14, 22, 15, 12, 14, 14, 13, 15, 16, 16, 20, 22]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.sheet_view.showGridLines = False

    def add_ranking_sheet(title, rows, table_name):
        target = workbook.create_sheet(title)
        target.append(headers)
        for row in rows:
            target.append(_excel_safe_row([
                row["rank"], row["region"], row["area"], row.get("code", ""), row["name"], row["total"], row["high"],
                row["medium"], row["low"], row["score_total"], row["max_score"], row["avg_score"],
                row["risk_label"], row["verified"], row["unverified"],
            ]))
        for cell in target[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="145DA0")
        if rows:
            table = Table(displayName=table_name, ref=f"A1:O{target.max_row}")
            table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
            target.add_table(table)
        else:
            target.auto_filter.ref = "A1:O1"
        target.freeze_panes = "A2"
        target.sheet_view.showGridLines = False
        for index, width in enumerate(widths, start=1):
            target.column_dimensions[get_column_letter(index)].width = width

    add_ranking_sheet("10 Terparah", location_rows[:10], "RankingTerparahFEWS")
    add_ranking_sheet("10 Terendah", list(reversed(location_rows[-10:])), "RankingTerendahFEWS")

    detail = workbook.create_sheet("Detail Temuan")
    detail_headers = [
        "ID Unix", "Tanggal", "Wilayah", "Area", "Kode Lokasi", "Lokasi", "Kesalahan",
        "Jumlah Kesalahan", "Skor", "Tipe Data", "Status Verifikasi",
    ]
    detail.append(detail_headers)
    for result in detail_rows:
        detail.append(_detail_row(result))
    for cell in detail[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="145DA0")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    if detail_rows:
        table = Table(displayName="DetailTemuanFEWS", ref=f"A1:K{detail.max_row}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        detail.add_table(table)
    else:
        detail.auto_filter.ref = "A1:K1"
    detail.freeze_panes = "A2"
    for index, width in enumerate([18, 13, 22, 24, 14, 22, 48, 18, 10, 14, 22], start=1):
        detail.column_dimensions[get_column_letter(index)].width = width
    detail.sheet_view.showGridLines = False

    trend_sheet = workbook.create_sheet("Tren Periode")
    trend_sheet.append(["Periode", "Total Skor"])
    values = (trend.get("series") or [{"values": []}])[0].get("values", [])
    for label, value in zip(trend.get("labels", []), values):
        trend_sheet.append(_excel_safe_row([label, value]))
    if trend_sheet.max_row > 1:
        chart = LineChart()
        chart.title = "Perbandingan Skor per Periode"
        chart.y_axis.title = "Total Skor"
        chart.x_axis.title = "Periode"
        chart.add_data(Reference(trend_sheet, min_col=2, min_row=1, max_row=trend_sheet.max_row), titles_from_data=True)
        chart.set_categories(Reference(trend_sheet, min_col=1, min_row=2, max_row=trend_sheet.max_row))
        chart.series[0].cat = AxDataSource(strRef=StrRef(f"'Tren Periode'!$A$2:$A${trend_sheet.max_row}"))
        chart.series[0].graphicalProperties.line.solidFill = "1677D2"
        chart.series[0].graphicalProperties.line.width = 28575
        chart.series[0].marker.symbol = "circle"
        chart.series[0].marker.size = 6
        chart.series[0].smooth = False
        chart.legend = None
        chart.x_axis.axPos = "b"
        chart.y_axis.axPos = "l"
        chart.height = 8
        chart.width = 16
        trend_sheet.add_chart(chart, "D2")

    indicator_sheet = workbook.create_sheet("Grafik Indikator")
    indicator_sheet.append(["Indikator", "Jumlah Kesalahan", "Total Skor"])
    for row in indicator_rows:
        indicator_sheet.append(_excel_safe_row([row["name"], row["value"], row.get("score", 0)]))
    if indicator_sheet.max_row > 1:
        chart = BarChart()
        chart.type = "bar"
        chart.title = "Kesalahan per Indikator"
        chart.x_axis.title = "Jumlah"
        chart.add_data(Reference(indicator_sheet, min_col=2, min_row=1, max_row=indicator_sheet.max_row), titles_from_data=True)
        chart.set_categories(Reference(indicator_sheet, min_col=1, min_row=2, max_row=indicator_sheet.max_row))
        chart.series[0].cat = AxDataSource(strRef=StrRef(f"'Grafik Indikator'!$A$2:$A${indicator_sheet.max_row}"))
        chart.series[0].graphicalProperties.solidFill = "1677D2"
        chart.series[0].graphicalProperties.line.solidFill = "1677D2"
        chart.dLbls = DataLabelList()
        chart.dLbls.showVal = True
        chart.legend = None
        chart.height = 9
        chart.width = 17
        indicator_sheet.add_chart(chart, "E2")
    indicator_sheet.column_dimensions["A"].width = 52

    location_sheet = workbook.create_sheet("Grafik Lokasi")
    location_sheet.append(["Lokasi", "Total Skor", "Jumlah Temuan"])
    for row in location_rows[:10]:
        label = f"{row.get('code')} - {row['name']}" if row.get("code") else row["name"]
        location_sheet.append(_excel_safe_row([label, row["score_total"], row["total"]]))
    if location_sheet.max_row > 1:
        chart = BarChart()
        chart.type = "bar"
        chart.title = "10 Lokasi dengan Skor Tertinggi"
        chart.x_axis.title = "Total Skor"
        chart.add_data(Reference(location_sheet, min_col=2, min_row=1, max_row=location_sheet.max_row), titles_from_data=True)
        chart.set_categories(Reference(location_sheet, min_col=1, min_row=2, max_row=location_sheet.max_row))
        chart.series[0].cat = AxDataSource(strRef=StrRef(f"'Grafik Lokasi'!$A$2:$A${location_sheet.max_row}"))
        chart.series[0].graphicalProperties.solidFill = "E5484D"
        chart.series[0].graphicalProperties.line.solidFill = "E5484D"
        chart.dLbls = DataLabelList()
        chart.dLbls.showVal = True
        chart.legend = None
        chart.height = 9
        chart.width = 17
        location_sheet.add_chart(chart, "E2")
    location_sheet.column_dimensions["A"].width = 30

    for target in [trend_sheet, indicator_sheet, location_sheet]:
        for cell in target[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="145DA0")
        target.freeze_panes = "A2"
        target.sheet_view.showGridLines = False

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_pdf_report(
    transactions,
    region: str = "",
    *,
    filters=None,
    trend=None,
    indicator_rows=None,
    location_rows=None,
):
    from reportlab.lib import colors
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    from reportlab.graphics.charts.linecharts import HorizontalLineChart
    from reportlab.graphics.shapes import Drawing, String
    from reportlab.lib.enums import TA_CENTER
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    from xml.sax.saxutils import escape

    filters = filters or {}
    trend = trend or {"labels": [], "series": [], "analysis": {}}
    indicator_rows = indicator_rows or []
    location_rows = location_rows or []
    output = BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=14 * mm,
        bottomMargin=14 * mm,
    )
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="FEWSSection", parent=styles["Heading2"], textColor=colors.HexColor("#0B2B4C"), spaceBefore=8, spaceAfter=7))
    styles.add(ParagraphStyle(name="FEWSSmall", parent=styles["BodyText"], fontSize=7, leading=9))
    styles.add(ParagraphStyle(name="FEWSCenter", parent=styles["FEWSSmall"], alignment=TA_CENTER))
    title = f"Laporan FEWS Wilayah {escape(region)}" if region else "Laporan FEWS Per Lokasi"
    period_label = filters.get("week") if filters.get("period_type") == "mingguan" else filters.get("month")
    filter_label = " | ".join([
        f"Periode: {period_label or 'Semua'} ({filters.get('period_type', 'bulanan')})",
        f"Area: {filters.get('area') or 'Semua'}",
        f"Lokasi: {filters.get('location') or 'Semua'}",
        f"Kesalahan: {filters.get('indicator') or 'Semua'}",
        f"Verifikasi: {filters.get('verification') or 'Semua'}",
    ])
    elements = [
        Paragraph(title, styles["Title"]),
        Paragraph(escape(filter_label), styles["FEWSSmall"]),
        Spacer(1, 8),
    ]

    common_table_style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#123b5d")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#A8BAC8")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.HexColor("#eef4f8")]),
    ])

    ranking_headers = ["Rank", "Wilayah", "Area", "Kode", "Lokasi", "Temuan", "Skor", "Skor Max", "Risiko", "Sudah", "Belum"]
    def ranking_table(rows):
        data = [ranking_headers]
        for row in rows:
            data.append([
                row["rank"], row["region"], row["area"], row.get("code", ""), row["name"], row["total"], row["score_total"],
                row["max_score"], row["risk_label"], row["verified"], row["unverified"],
            ])
        table = Table(data, repeatRows=1, colWidths=[12*mm, 27*mm, 30*mm, 14*mm, 30*mm, 15*mm, 15*mm, 17*mm, 17*mm, 15*mm, 15*mm])
        table.setStyle(common_table_style)
        return table

    elements.extend([
        Paragraph("10 Risiko Terparah", styles["FEWSSection"]),
        ranking_table(location_rows[:10]),
        Paragraph("10 Risiko Terendah", styles["FEWSSection"]),
        ranking_table(list(reversed(location_rows[-10:]))),
        PageBreak(),
        Paragraph("Visualisasi FEWS", styles["FEWSSection"]),
    ])

    chart_width, chart_height = 360, 180
    trend_values = (trend.get("series") or [{"values": []}])[0].get("values", [])
    if trend_values:
        drawing = Drawing(chart_width, chart_height + 25)
        drawing.add(String(8, chart_height + 10, "Perbandingan Skor per Periode", fontName="Helvetica-Bold", fontSize=10, fillColor=colors.HexColor("#0B2B4C")))
        chart = HorizontalLineChart()
        chart.x, chart.y, chart.width, chart.height = 42, 35, 300, 125
        chart.data = [trend_values]
        chart.categoryAxis.categoryNames = trend.get("labels", [])
        chart.categoryAxis.labels.fontSize = 7
        chart.valueAxis.valueMin = 0
        chart.lines[0].strokeColor = colors.HexColor("#1677D2")
        chart.lines[0].strokeWidth = 2
        drawing.add(chart)
        elements.append(drawing)
        analysis = trend.get("analysis") or {}
        if analysis.get("headline"):
            elements.append(Paragraph(f"<b>Analisis:</b> {escape(analysis['headline'])}. {escape(analysis.get('recommendation', ''))}", styles["BodyText"]))

    def bar_drawing(title_text, labels, values, fill):
        drawing = Drawing(chart_width, chart_height + 25)
        drawing.add(String(8, chart_height + 10, title_text, fontName="Helvetica-Bold", fontSize=10, fillColor=colors.HexColor("#0B2B4C")))
        chart = VerticalBarChart()
        chart.x, chart.y, chart.width, chart.height = 42, 48, 300, 110
        chart.data = [values]
        chart.categoryAxis.categoryNames = [label[:13] for label in labels]
        chart.categoryAxis.labels.fontSize = 6
        chart.categoryAxis.labels.angle = 25
        chart.valueAxis.valueMin = 0
        chart.bars[0].fillColor = colors.HexColor(fill)
        drawing.add(chart)
        return drawing

    if indicator_rows:
        elements.append(bar_drawing(
            "Jumlah Kesalahan per Indikator",
            [row["name"] for row in indicator_rows],
            [row["value"] for row in indicator_rows],
            "#1677D2",
        ))
    if location_rows:
        elements.append(bar_drawing(
            "Total Skor per Lokasi (10 Terparah)",
            [f"{row.get('code', '')} {row['name']}".strip() for row in location_rows[:10]],
            [row["score_total"] for row in location_rows[:10]],
            "#E5484D",
        ))

    elements.extend([PageBreak(), Paragraph("Detail Temuan", styles["FEWSSection"])])
    detail_data = [["ID Unix", "Tanggal", "Wilayah", "Area", "Kode", "Lokasi", "Kesalahan", "Jumlah", "Skor", "Verifikasi"]]
    for result in transactions:
        row = _detail_row(result)
        detail_data.append([
            Paragraph(escape(str(row[0])), styles["FEWSSmall"]), row[1], row[2], row[3], row[4], row[5],
            Paragraph(escape(str(row[6])), styles["FEWSSmall"]), row[7], row[8], row[10],
        ])
    detail_table = Table(
        detail_data,
        repeatRows=1,
        colWidths=[23*mm, 17*mm, 25*mm, 27*mm, 13*mm, 27*mm, 55*mm, 13*mm, 12*mm, 28*mm],
    )
    detail_table.setStyle(common_table_style)
    elements.append(detail_table)

    def page_number(canvas, document):
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(colors.HexColor("#5F7285"))
        canvas.drawRightString(landscape(A4)[0] - 12 * mm, 7 * mm, f"FEWS | Halaman {document.page}")
        canvas.restoreState()

    doc.build(elements, onFirstPage=page_number, onLaterPages=page_number)
    return output.getvalue()
